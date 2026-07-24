from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook

from app.modules.leads.opt.parser import parse_application_workbook


def _build_opt_xlsx(rows: list[tuple]) -> bytes:
    """Minimal OPT заявка: headers row 1-3, data from row 4. Cols B-E."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Заявка"
    ws["B3"] = "ИНН продавца"
    ws["C3"] = "ИНН покупателя"
    ws["D3"] = "Дата"
    ws["E3"] = "Сумма"
    for idx, row in enumerate(rows, start=4):
        for col, value in enumerate(row, start=2):
            ws.cell(idx, col, value)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_skips_mid_file_gaps_and_reads_tail() -> None:
    """Regression: empty rows in the middle used to truncate the file."""
    content = _build_opt_xlsx(
        [
            ("7708721010", "9102191311", date(2026, 6, 10), Decimal("421000.00")),
            ("7708721010", "9102191311", date(2026, 6, 20), Decimal("155000.00")),
            (None, None, None, None),
            (None, None, None, None),
            ("7708721010", None, None, None),  # junk spacer with only partial junk
            (None, None, None, None),
            ("7708721010", "9102191311", date(2026, 6, 16), Decimal("132600.00")),
            ("7708721010", "9102191311", date(2026, 6, 23), Decimal("568365.00")),
            ("Итого:", None, None, Decimal("1276965.00")),
        ],
    )
    parsed = parse_application_workbook(content)
    assert parsed.buyer_inn == "9102191311"
    assert len(parsed.lines) == 4
    assert parsed.lines[-1].amount == Decimal("568365.00")
    assert parsed.lines[-2].document_date == date(2026, 6, 16)
    assert sum(line.amount for line in parsed.lines) == Decimal("1276965.00")


def test_parse_carries_forward_inn_when_blank_on_tail_rows() -> None:
    content = _build_opt_xlsx(
        [
            ("9718148521", "9725198516", date(2026, 4, 30), Decimal("1285600.00")),
            ("9718148521", "9725198516", date(2026, 5, 31), Decimal("1016200.00")),
            (None, None, date(2026, 6, 16), Decimal("132600.00")),
            (None, None, date(2026, 6, 23), Decimal("568365.00")),
            ("Итого", None, None, None),
        ],
    )
    parsed = parse_application_workbook(content)
    assert len(parsed.lines) == 4
    assert parsed.lines[2].supplier_inn == "9718148521"
    assert parsed.lines[2].buyer_inn == "9725198516"
    assert parsed.lines[3].amount == Decimal("568365.00")
