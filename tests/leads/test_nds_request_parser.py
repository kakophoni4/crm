from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.modules.leads.opt.nds_request_parser import (
    looks_like_nds_request,
    parse_nds_request_workbook,
)


def test_parse_sample_nds_request_file() -> None:
    path = Path(__file__).resolve().parents[1] / "63. ЗАПРОС НДС 2 КВ. 2026Г..xlsx"
    if not path.exists():
        # Sample workbook may be absent in CI — skip lightly.
        return
    content = path.read_bytes()
    assert looks_like_nds_request(content)
    result = parse_nds_request_workbook(content)
    assert result.matched
    assert result.form_kind == "nds_request"
    assert result.application is not None
    assert result.application.buyer_inn == "6321443710"
    assert len(result.application.lines) >= 40
    assert result.application.lines[0].supplier_inn == "9704271074"
    assert result.application.lines[0].amount > 0


def test_parse_partner_forma_zayavki_headers() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Лист_1"
    ws.append(
        [
            "Ваши контактные данные",
            "Наименование покупателя",
            "ИНН покупателя",
            "КПП покупателя",
            "Вид товара",
            "Дата (дд.мм.гг)",
            "Сумма (в т.ч. НДС)",
            "Сумма НДС",
            "ИНН организации",
            "Ставка НДС",
        ],
    )
    ws.append(
        [
            "tg",
            "ООО Тест",
            "7701234567",
            "770101001",
            "услуги",
            date(2026, 4, 15),
            1_000_000,
            166_666.67,
            "9704271074",
            20,
        ],
    )
    buf = BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    assert looks_like_nds_request(content)
    result = parse_nds_request_workbook(content)
    assert result.matched
    assert result.form_kind == "partner_forma"
    assert result.application is not None
    assert result.application.buyer_inn == "7701234567"
    assert len(result.application.lines) == 1
    assert result.application.lines[0].supplier_inn == "9704271074"
    assert result.application.lines[0].amount == 1_000_000


def test_reject_crm_registry_headers() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "TDSheet"
    ws.append(
        [
            "№ документа",
            "Дата документа",
            "Покупатель",
            "ИНН покупателя",
            "Поставщик",
            "ИНН поставщика",
            "Сумма",
            "Сумма НДС",
            "Сумма без НДС",
        ],
    )
    ws.append(["1", date(2026, 4, 1), "ООО", "7701234567", "ООО2", "9704271074", 100, 20, 80])
    buf = BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    result = parse_nds_request_workbook(content)
    assert result.matched is False


def test_reject_random_bytes() -> None:
    assert not looks_like_nds_request(b"not-an-excel")
    result = parse_nds_request_workbook(b"not-an-excel")
    assert result.matched is False


def test_parse_partner_forma_xls() -> None:
    xlwt = pytest.importorskip("xlwt")
    book = xlwt.Workbook()
    sheet = book.add_sheet("Лист_1")
    headers = [
        "Ваши контактные данные",
        "Наименование покупателя",
        "ИНН покупателя",
        "КПП покупателя",
        "Вид товара",
        "Дата (дд.мм.гг)",
        "Сумма (в т.ч. НДС)",
        "Сумма НДС",
        "ИНН организации",
        "Ставка НДС",
    ]
    for col, text in enumerate(headers):
        sheet.write(0, col, text)
    sheet.write(1, 0, "tg")
    sheet.write(1, 1, "ООО Тест")
    sheet.write(1, 2, "7701234567")
    sheet.write(1, 3, "770101001")
    sheet.write(1, 4, "услуги")
    sheet.write(1, 5, "15.04.2026")
    sheet.write(1, 6, 1_000_000)
    sheet.write(1, 7, 166_666.67)
    sheet.write(1, 8, "9704271074")
    sheet.write(1, 9, 22)
    buf = BytesIO()
    book.save(buf)
    content = buf.getvalue()
    assert content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    assert looks_like_nds_request(content)
    result = parse_nds_request_workbook(content)
    assert result.matched
    assert result.form_kind == "partner_forma"
    assert result.application is not None
    assert result.application.buyer_inn == "7701234567"
    assert len(result.application.lines) == 1
    assert result.application.lines[0].amount == 1_000_000


def test_parse_park_zapros_xls_headers() -> None:
    """Alternate layout: ИНН нашей/вашей компании + сумма сделок."""
    xlwt = pytest.importorskip("xlwt")
    book = xlwt.Workbook()
    sheet = book.add_sheet("Sheet1")
    # title rows like real files
    sheet.write(0, 0, "Запрос")
    sheet.write(1, 0, "")
    headers = [
        "Требуемый ОКВЭД",
        "ИНН нашей компании",
        "ИНН Вашей компании",
        "Дата с/ф\n(дд. мм. гггг)",
        "Сумма сделок\n(НДС в том числе)",
    ]
    for col, text in enumerate(headers):
        sheet.write(2, col, text)
    sheet.write(3, 0, "62.01")
    sheet.write(3, 1, "9704271074")
    sheet.write(3, 2, "3525479879")
    sheet.write(3, 3, "10.04.2026")
    sheet.write(3, 4, 500_000)
    buf = BytesIO()
    book.save(buf)
    result = parse_nds_request_workbook(buf.getvalue())
    assert result.matched
    assert result.form_kind == "partner_forma"
    assert result.application is not None
    assert result.application.buyer_inn == "3525479879"
    assert result.application.lines[0].supplier_inn == "9704271074"
    assert result.application.lines[0].amount == 500_000


def test_parse_easy_goldman_zapros_headers() -> None:
    """Easy Goldman: Компании-продавца/покупатели + Сумма покупок."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Заявка", "Ставка НДС:", 20, "%"])
    ws.append(
        [
            "ИНН Компании-продавца (Наши компании)",
            "ИНН Компании-покупатели (Ваши компании)",
            "Дата с/ф (дд. мм. гггг)",
            "Сумма покупок (НДС в том числе)",
        ],
    )
    ws.append(["9709103059", "6678133104", "09.07.2024", 489_000])
    ws.append(["9709103059", "6678133104", "10.07.2024", 99_000])
    buf = BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    assert looks_like_nds_request(content)
    result = parse_nds_request_workbook(content)
    assert result.matched
    assert result.form_kind == "partner_forma"
    assert result.application is not None
    assert result.application.buyer_inn == "6678133104"
    assert len(result.application.lines) == 2
    assert result.application.lines[0].supplier_inn == "9709103059"
    assert result.application.lines[0].amount == 489_000
    assert result.application.lines[1].amount == 99_000


def test_parse_tavrida_summa_pokupki_po_sf() -> None:
    """Таврида / Заявка новая: Сумма покупки по СФ, в т.ч. НДС."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append([None, None, None])
    ws.append([None, None, "Ставка НДС :", 22, "%"])
    ws.append(
        [
            None,
            "ИНН ПРОДАВЦА\n",
            "ИНН ПОКУПАТЕЛЯ\n",
            "Дата СФ\n(в виде дд. мм. гггг)",
            "Сумма покупки по СФ, в т.ч. НДС\n",
        ],
    )
    ws.append([None, "9709103059", "6678133104", "09.07.2024", 489_000])
    ws.append([None, "9709103059", "6678133104", "10.07.2024", 99_000])
    buf = BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    assert looks_like_nds_request(content)
    result = parse_nds_request_workbook(content)
    assert result.matched
    assert result.form_kind == "nds_request"
    assert result.application is not None
    assert result.application.buyer_inn == "6678133104"
    assert len(result.application.lines) == 2
    assert result.application.lines[0].amount == 489_000


def test_partner_forma_header_below_row_8() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    for _ in range(10):
        ws.append(["title", "banner"])
    ws.append(
        [
            "Ваши контактные данные (ник теле, номер телефона): XXX",
            "Наименование покупателя",
            "ИНН Покупателя",
            "КПП Покупателя",
            "Вид товара/работы/услуги и ОКВЭД",
            "Дата (дд.мм.гг)",
            "Сумма (в т.ч. НДС)",
            "Сумма НДС",
            "ИНН Организации",
            "Ставка НДС",
        ],
    )
    ws.append(
        [
            "tg",
            'ООО "ДОМ ЗАПЧАСТЕЙ"',
            "3 525 462 748",
            "3 525 010 01",
            "услуги",
            "08.04.2026",
            85360,
            "15392,79",
            "9704271074",
            22,
        ],
    )
    buf = BytesIO()
    wb.save(buf)
    result = parse_nds_request_workbook(buf.getvalue())
    assert result.matched
    assert result.form_kind == "partner_forma"
    assert result.application is not None
    assert result.application.buyer_inn == "3525462748"
