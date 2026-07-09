from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.modules.leads.opt.contact_buyer import normalize_inn, parse_decimal, parse_excel_date

PATH = Path(__file__).resolve().parent / "Форма заявки.xlsx"

inn_re = re.compile(r"^\d{10}(\d{2})?$")

wb = load_workbook(PATH, data_only=True)
ws = wb.active
print("file:", PATH)
print("title:", ws.title)
print("max_row:", ws.max_row, "max_col:", ws.max_column)
print("\n--- A:F rows 1..6 ---")
for r in range(1, 7):
    print(r, [ws.cell(r, c).value for c in range(1, 7)])

print("\n--- B:E rows 4..30 (parser view) ---")
buyer_inn: str | None = None
valid_rows: list[int] = []
parser_lines: list[int] = []

for row_idx in range(4, min(ws.max_row or 0, 30) + 1):
    raw_b = ws.cell(row_idx, 2).value
    raw_c = ws.cell(row_idx, 3).value
    raw_d = ws.cell(row_idx, 4).value
    raw_e = ws.cell(row_idx, 5).value

    supplier_inn = normalize_inn(raw_b)
    row_buyer_inn = normalize_inn(raw_c)
    document_date = parse_excel_date(raw_d)
    amount = parse_decimal(raw_e)

    would_stop = False
    if supplier_inn is None and row_buyer_inn is None:
        would_stop = bool(parser_lines)
    elif supplier_inn is None or document_date is None or amount is None or amount <= 0:
        would_stop = bool(parser_lines)

    if row_buyer_inn:
        buyer_inn = row_buyer_inn

    row_ok = (
        supplier_inn is not None
        and document_date is not None
        and amount is not None
        and amount > 0
        and buyer_inn is not None
    )
    if row_ok:
        valid_rows.append(row_idx)
        parser_lines.append(row_idx)

    print(
        row_idx,
        [raw_b, raw_c, raw_d, raw_e],
        "=>",
        {
            "supplier_inn": supplier_inn,
            "buyer_inn": row_buyer_inn,
            "date": document_date,
            "amount": amount,
            "buyer_seen": buyer_inn,
            "row_ok": row_ok,
            "parser_would_stop_here": would_stop,
        },
    )

print("\nVALID_ROWS (strict):", valid_rows)
print("PARSER_WOULD_ACCEPT:", len(parser_lines), "lines, buyer_inn=", buyer_inn)
