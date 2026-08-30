from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.leads.opt.contact_buyer import normalize_inn, parse_decimal, parse_excel_date
from app.shared.exceptions import ValidationError


@dataclass(frozen=True)
class ParsedApplicationLine:
    supplier_inn: str
    buyer_inn: str
    document_date: date
    amount: Decimal
    supplier_name: str | None = None
    supplier_kpp: str | None = None


@dataclass(frozen=True)
class ParsedApplication:
    buyer_inn: str
    lines: list[ParsedApplicationLine]
    buyer_name: str | None = None
    buyer_kpp: str | None = None


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace("ё", "е")


def _is_total_or_stop_row(worksheet: Worksheet, row_idx: int) -> bool:
    """Stop markers: «Итого», totals — must not become data lines."""
    for col in range(1, 8):
        text = _cell_text(worksheet.cell(row_idx, col).value)
        if not text:
            continue
        compact = text.replace(":", "").replace(" ", "")
        if compact.startswith("итого") or compact in {"total", "sum", "сумма"}:
            return True
    return False


def _row_has_more_data_ahead(
    worksheet: Worksheet,
    row_idx: int,
    max_row: int,
    *,
    peek: int = 40,
) -> bool:
    """True if a later row still looks like an invoice line (gap in the middle)."""
    end = min(row_idx + peek, max_row)
    for peek_idx in range(row_idx + 1, end + 1):
        if _is_total_or_stop_row(worksheet, peek_idx):
            return False
        supplier_inn = normalize_inn(worksheet.cell(peek_idx, 2).value)
        document_date = parse_excel_date(worksheet.cell(peek_idx, 4).value)
        amount = parse_decimal(worksheet.cell(peek_idx, 5).value)
        if supplier_inn and document_date is not None and amount is not None and amount > 0:
            return True
        # Date+amount with blank INNs — still a data row (INN will be carried forward).
        if document_date is not None and amount is not None and amount > 0:
            return True
    return False


def parse_application_workbook(content: bytes) -> ParsedApplication:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValidationError(message="Не удалось прочитать Excel-файл заявки") from exc

    worksheet = workbook.active
    lines: list[ParsedApplicationLine] = []
    buyer_inn: str | None = None
    last_supplier_inn: str | None = None
    max_row = int(worksheet.max_row or 0)

    for row_idx in range(4, max_row + 1):
        if _is_total_or_stop_row(worksheet, row_idx):
            break

        supplier_inn = normalize_inn(worksheet.cell(row_idx, 2).value)
        row_buyer_inn = normalize_inn(worksheet.cell(row_idx, 3).value)
        document_date = parse_excel_date(worksheet.cell(row_idx, 4).value)
        amount = parse_decimal(worksheet.cell(row_idx, 5).value)

        # Blank / spacer rows: skip, do NOT stop — clients often leave gaps mid-file.
        if supplier_inn is None and row_buyer_inn is None and document_date is None and amount is None:
            if lines and not _row_has_more_data_ahead(worksheet, row_idx, max_row):
                break
            continue

        # Carry forward INN when date+amount present but seller/buyer cells blank.
        if supplier_inn is None and last_supplier_inn is not None:
            supplier_inn = last_supplier_inn
        if row_buyer_inn is None and buyer_inn is not None:
            row_buyer_inn = buyer_inn

        if supplier_inn is None or document_date is None or amount is None or amount <= 0:
            # Incomplete junk row: skip without aborting the rest of the sheet.
            continue

        if row_buyer_inn:
            buyer_inn = row_buyer_inn
        elif buyer_inn is None:
            raise ValidationError(message="В заявке не найден ИНН покупателя")

        last_supplier_inn = supplier_inn
        lines.append(
            ParsedApplicationLine(
                supplier_inn=supplier_inn,
                buyer_inn=buyer_inn or row_buyer_inn or "",
                document_date=document_date,
                amount=amount,
            ),
        )

    if not lines:
        raise ValidationError(message="В файле заявки не найдено строк с операциями")
    if buyer_inn is None:
        buyer_inn = lines[0].buyer_inn
    if not buyer_inn:
        raise ValidationError(message="ИНН покупателя обязателен")

    return ParsedApplication(buyer_inn=buyer_inn, lines=lines)
