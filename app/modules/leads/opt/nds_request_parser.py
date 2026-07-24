"""Detect and parse partner «ЗАПРОС НДС» / Forma_zayavki Excel by content.

Two partner layouts (OPT upload format is intentionally NOT handled here):

1) nds_request — sheet «Заявка на НДС»
   ИНН покупателя | Стоимость покупки | ИНН продавца | дата счета-фактуры

2) partner_forma — Forma_zayavki
   ИНН покупателя | Сумма (в т.ч. НДС) | ИНН организации | дата (дд.мм.гг)

3) park / Easy Goldman zapros
   ИНН Компании-продавца (Наши) | ИНН Компании-покупатели (Ваши)
   | Дата с/ф | Сумма покупок (НДС в том числе)
   (also: ИНН нашей/вашей компании + Сумма сделок)

4) «Запрос на подписание» (Улитин-style)
   ИНН покупателя | Дата с-ф | Сумма с-ф | ИНН продавца

Supports .xlsx (openpyxl) and legacy .xls (xlrd).
CRM registry exports (№ документа / поставщик / сумма без НДС) are rejected.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Literal

from openpyxl import load_workbook

from app.modules.leads.opt.contact_buyer import normalize_inn, parse_decimal, parse_excel_date
from app.modules.leads.opt.parser import ParsedApplication, ParsedApplicationLine

FormKind = Literal["nds_request", "partner_forma"]

_XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


@dataclass(frozen=True)
class NdsRequestParseResult:
    matched: bool
    sheet_name: str | None = None
    application: ParsedApplication | None = None
    reason: str | None = None
    form_kind: FormKind | None = None


@dataclass(frozen=True)
class _SheetMatrix:
    title: str
    rows: list[list[object]]  # 0-based rows; cells may be shorter than max col


_NDS_MARKERS = (
    "инн покупателя",
    "стоимость покупки",
    "инн продавца",
)

_PARTNER_FORMA_MARKERS = (
    "инн покупателя",
    "сумма (в т.ч. ндс)",
    "инн организации",
)

_HEADER_SCAN_ROWS = 20
_HEADER_SCAN_COLS = 24


def _is_xls(content: bytes) -> bool:
    return len(content) >= 8 and content[:8] == _XLS_MAGIC


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def _blob(values: list[object] | tuple[object, ...]) -> str:
    return " ".join(_cell_text(v) for v in values if v is not None)


def _is_crm_registry_blob(blob: str) -> bool:
    return (
        "сумма без ндс" in blob
        and ("№ документа" in blob or "номер документа" in blob)
        and "поставщик" in blob
    )


def _header_map(row_values: list[object]) -> dict[str, int]:
    """Map logical fields → 1-based column index from a header row."""
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(row_values, start=1):
        text = _cell_text(raw)
        if not text:
            continue
        # Buyer first: «покупателя/покупатели», «Вашей/Ваши компании»
        if "инн" in text and "покупател" in text:
            mapping["buyer_inn"] = idx
        elif "инн вашей" in text or "инн вашей компании" in text:
            mapping["buyer_inn"] = idx
        elif "ваши компани" in text and "инн" in text:
            # Easy Goldman: «ИНН Компании-покупатели (Ваши компании)»
            mapping["buyer_inn"] = idx
        elif "инн нашей" in text or "инн нашей компании" in text:
            mapping["supplier_inn"] = idx
        elif "наши компани" in text and "инн" in text:
            # Easy Goldman: «ИНН Компании-продавца (Наши компании)»
            mapping["supplier_inn"] = idx
        elif "инн организации" in text:
            mapping["supplier_inn"] = idx
        elif "инн продавца" in text or ("инн" in text and "продав" in text):
            mapping["supplier_inn"] = idx
        elif "стоимость покупки" in text:
            mapping["amount"] = idx
        elif "сумма с-ф" in text or "сумма сф" in text or text.startswith("сумма с-ф"):
            # «Запрос на подписание»: Сумма с-ф
            mapping["amount"] = idx
        elif "сумма покупки" in text and "ндс" in text:
            # Таврида / «Заявка новая»: Сумма покупки по СФ, в т.ч. НДС
            mapping["amount"] = idx
        elif "сумма покупок" in text and "ндс" in text:
            mapping["amount"] = idx
        elif "сумма сделок" in text and "ндс" in text:
            mapping["amount"] = idx
        elif "сумма (в т.ч. ндс)" in text or "сумма в т.ч. ндс" in text:
            mapping["amount"] = idx
        elif "в т.ч. ндс" in text and "сумма" in text and "сумма ндс" not in text:
            mapping["amount"] = idx
        elif text == "сумма ндс" or text.startswith("сумма ндс "):
            continue
        elif (
            "дата с/ф" in text
            or "дата с-ф" in text
            or "дата сф" in text
            or "дата с / ф" in text
        ):
            mapping["document_date"] = idx
        elif "дата счета" in text or "дата счёта" in text or "дата счет" in text:
            mapping["document_date"] = idx
        elif (
            text.startswith("дата (дд")
            or text.startswith("дата(дд")
            or text == "дата"
            or text.startswith("дата реализации")
        ):
            if "document_date" not in mapping or "реализац" not in text:
                mapping["document_date"] = idx
        elif text.startswith("наименование покупателя"):
            mapping["buyer_name"] = idx
        elif "наименование продавца" in text or "наименование организации" in text:
            mapping["supplier_name"] = idx
    return mapping


def _is_park_zapros_blob(blob: str) -> bool:
    """Park / Easy Goldman «Заявка» layouts (наши/ваши + сумма покупок/сделок)."""
    has_parties = (
        ("инн вашей" in blob and "инн нашей" in blob)
        or ("ваши компани" in blob and "наши компани" in blob)
        or ("покупател" in blob and "продав" in blob and "инн" in blob)
    )
    has_amount = (
        "сумма сделок" in blob
        or "сумма покупок" in blob
        or "сумма покупки" in blob
        or "стоимость покупки" in blob
        or ("сумма" in blob and "в т.ч" in blob and "ндс" in blob)
    )
    return has_parties and has_amount and "дата" in blob


def _detect_layout(blob: str, mapping: dict[str, int]) -> FormKind | None:
    required = {"buyer_inn", "supplier_inn", "amount", "document_date"}
    if not required.issubset(mapping):
        return None
    if _is_crm_registry_blob(blob):
        return None
    if all(m in blob for m in _NDS_MARKERS):
        return "nds_request"
    if _is_park_zapros_blob(blob):
        return "partner_forma"
    if "инн организации" in blob and (
        "сумма (в т.ч. ндс)" in blob
        or "сумма в т.ч. ндс" in blob
        or ("сумма" in blob and "в т.ч" in blob and "ндс" in blob)
    ):
        return "partner_forma"
    if all(m in blob for m in _PARTNER_FORMA_MARKERS):
        return "partner_forma"
    if "инн покупателя" in blob and "инн" in blob and (
        "стоимость" in blob
        or "сумма (в т.ч" in blob
        or "сумма покупок" in blob
        or "сумма покупки" in blob
        or "сумма с-ф" in blob
        or "сумма сф" in blob
        or ("сумма" in blob and "в т.ч" in blob and "ндс" in blob)
    ):
        if "инн продавца" in blob or "продав" in blob:
            return "nds_request"
        if "инн организации" in blob:
            return "partner_forma"
    return None


def _matrix_row(sheet: _SheetMatrix, row_idx: int, max_col: int) -> list[object]:
    """1-based row_idx → values padded/truncated to max_col."""
    if row_idx < 1 or row_idx > len(sheet.rows):
        return [None] * max_col
    raw = sheet.rows[row_idx - 1]
    out: list[object] = []
    for col in range(max_col):
        out.append(raw[col] if col < len(raw) else None)
    return out


def _cell(sheet: _SheetMatrix, row_idx: int, col_idx: int) -> object:
    """1-based row/col."""
    if row_idx < 1 or col_idx < 1:
        return None
    if row_idx > len(sheet.rows):
        return None
    row = sheet.rows[row_idx - 1]
    if col_idx > len(row):
        return None
    return row[col_idx - 1]


def _sheet_looks_like_partner(
    sheet: _SheetMatrix,
) -> tuple[int, dict[str, int], FormKind] | None:
    max_row = min(len(sheet.rows), _HEADER_SCAN_ROWS)
    if max_row < 1:
        return None
    max_col = 0
    for row in sheet.rows[:max_row]:
        max_col = max(max_col, len(row))
    max_col = min(max_col, _HEADER_SCAN_COLS)
    if max_col < 4:
        return None
    for row_idx in range(1, max_row + 1):
        values = _matrix_row(sheet, row_idx, max_col)
        blob = _blob(values)
        if not blob or _is_crm_registry_blob(blob):
            continue
        mapping = _header_map(values)
        kind = _detect_layout(blob, mapping)
        if kind is not None:
            return row_idx, mapping, kind
    return None


def _load_xlsx_sheets(content: bytes) -> list[_SheetMatrix]:
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheets: list[_SheetMatrix] = []
    try:
        for ws in workbook.worksheets:
            max_col = min(int(ws.max_column or 0), _HEADER_SCAN_COLS)
            max_row = int(ws.max_row or 0)
            rows = [
                [ws.cell(r, c).value for c in range(1, max_col + 1)]
                for r in range(1, max_row + 1)
            ]
            sheets.append(_SheetMatrix(title=ws.title, rows=rows))
    finally:
        workbook.close()
    return sheets


def _load_xls_sheets(content: bytes) -> list[_SheetMatrix]:
    try:
        import xlrd
        from xlrd import xldate_as_datetime
    except ImportError as exc:
        raise RuntimeError("xlrd_not_installed") from exc

    book = xlrd.open_workbook(file_contents=content, formatting_info=False)
    sheets: list[_SheetMatrix] = []
    try:
        for sheet in book.sheets():
            rows: list[list[object]] = []
            max_col = min(int(sheet.ncols), _HEADER_SCAN_COLS)
            for r in range(sheet.nrows):
                row_vals: list[object] = []
                for c in range(max_col):
                    cell = sheet.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_EMPTY:
                        row_vals.append(None)
                    elif cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt = xldate_as_datetime(cell.value, book.datemode)
                            row_vals.append(dt.date() if isinstance(dt, datetime) else dt)
                        except Exception:
                            row_vals.append(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        # Keep as number; normalize_inn/parse_decimal handle floats.
                        row_vals.append(cell.value)
                    else:
                        row_vals.append(cell.value)
                rows.append(row_vals)
            sheets.append(_SheetMatrix(title=sheet.name, rows=rows))
    finally:
        release = getattr(book, "release_resources", None)
        if callable(release):
            release()
    return sheets


def _load_sheets(content: bytes) -> list[_SheetMatrix]:
    if _is_xls(content):
        return _load_xls_sheets(content)
    return _load_xlsx_sheets(content)


def _blob_has_partner_headers(blob: str) -> bool:
    if not blob or _is_crm_registry_blob(blob):
        return False
    if all(m in blob for m in _NDS_MARKERS):
        return True
    if "инн покупателя" in blob and "инн организации" in blob and (
        "сумма (в т.ч. ндс)" in blob or "сумма в т.ч. ндс" in blob
    ):
        return True
    if _is_park_zapros_blob(blob):
        return True
    # Таврида / Заявка новая: ИНН продавца+покупателя + сумма покупки по СФ
    if (
        "инн" in blob
        and "покупател" in blob
        and "продав" in blob
        and ("сумма покупки" in blob or ("сумма" in blob and "в т.ч" in blob and "ндс" in blob))
        and "дата" in blob
    ):
        return True
    # «Запрос на подписание»: Дата с-ф + Сумма с-ф + ИНН покупателя/продавца
    if (
        "инн" in blob
        and "покупател" in blob
        and "продав" in blob
        and ("сумма с-ф" in blob or "сумма сф" in blob)
        and ("дата с-ф" in blob or "дата сф" in blob or "дата с/ф" in blob)
    ):
        return True
    return False


def _quick_has_partner_headers(content: bytes) -> bool:
    try:
        sheets = _load_sheets(content)
    except RuntimeError:
        return False
    except Exception:
        return False
    for sheet in sheets:
        max_row = min(len(sheet.rows), _HEADER_SCAN_ROWS)
        for row_idx in range(1, max_row + 1):
            values = _matrix_row(sheet, row_idx, _HEADER_SCAN_COLS)
            if _blob_has_partner_headers(_blob(values)):
                return True
    return False


def peek_workbook_headers(content: bytes, *, max_rows: int = 8) -> list[dict[str, object]]:
    """Return first non-empty rows per sheet (for SKIP diagnostics / audit)."""
    try:
        sheets = _load_sheets(content)
    except RuntimeError as exc:
        if "xlrd_not_installed" in str(exc):
            return [{"sheet": None, "reason": "xlrd_not_installed"}]
        return [{"sheet": None, "reason": str(exc)}]
    except Exception as exc:
        return [{"sheet": None, "reason": f"excel_open_failed: {exc}"}]

    out: list[dict[str, object]] = []
    for sheet in sheets:
        for row_idx in range(1, min(len(sheet.rows), max_rows) + 1):
            values = _matrix_row(sheet, row_idx, _HEADER_SCAN_COLS)
            cells = [str(v).strip() for v in values if v is not None and str(v).strip()]
            if len(cells) < 2:
                continue
            out.append(
                {
                    "sheet": sheet.title,
                    "row": row_idx,
                    "headers": cells[:16],
                    "blob": _blob(values)[:240],
                },
            )
            break
    return out


def looks_like_nds_request(content: bytes) -> bool:
    return _quick_has_partner_headers(content)


def _parse_sheet(
    sheet: _SheetMatrix,
    header_row: int,
    cols: dict[str, int],
    form_kind: FormKind,
) -> NdsRequestParseResult:
    lines: list[ParsedApplicationLine] = []
    buyer_inn: str | None = None
    max_row = len(sheet.rows)
    for row_idx in range(header_row + 1, max_row + 1):
        supplier_inn = normalize_inn(_cell(sheet, row_idx, cols["supplier_inn"]))
        row_buyer = normalize_inn(_cell(sheet, row_idx, cols["buyer_inn"]))
        document_date = parse_excel_date(_cell(sheet, row_idx, cols["document_date"]))
        amount = parse_decimal(_cell(sheet, row_idx, cols["amount"]))

        if supplier_inn is None and row_buyer is None and amount is None:
            if lines:
                empty_ahead = True
                for peek in range(row_idx + 1, min(row_idx + 3, max_row + 1)):
                    if normalize_inn(_cell(sheet, peek, cols["supplier_inn"])):
                        empty_ahead = False
                        break
                if empty_ahead:
                    break
            continue
        if supplier_inn is None or document_date is None or amount is None or amount <= 0:
            continue
        if row_buyer:
            buyer_inn = row_buyer
        if buyer_inn is None:
            continue
        lines.append(
            ParsedApplicationLine(
                supplier_inn=supplier_inn,
                buyer_inn=buyer_inn,
                document_date=document_date,
                amount=amount,
            ),
        )

    if not lines:
        return NdsRequestParseResult(
            matched=True,
            sheet_name=sheet.title,
            reason="no_data_rows",
            form_kind=form_kind,
        )
    assert buyer_inn is not None
    return NdsRequestParseResult(
        matched=True,
        sheet_name=sheet.title,
        application=ParsedApplication(buyer_inn=buyer_inn, lines=lines),
        form_kind=form_kind,
    )


def parse_nds_request_workbook(content: bytes) -> NdsRequestParseResult:
    if _is_xls(content):
        try:
            import xlrd  # noqa: F401
        except ImportError:
            return NdsRequestParseResult(matched=False, reason="xlrd_not_installed")

    try:
        sheets = _load_sheets(content)
    except RuntimeError as exc:
        if "xlrd_not_installed" in str(exc):
            return NdsRequestParseResult(matched=False, reason="xlrd_not_installed")
        return NdsRequestParseResult(matched=False, reason=str(exc))
    except Exception as exc:
        return NdsRequestParseResult(matched=False, reason=f"excel_open_failed: {exc}")

    if not any(
        _blob_has_partner_headers(_blob(_matrix_row(s, r, _HEADER_SCAN_COLS)))
        for s in sheets
        for r in range(1, min(len(s.rows), _HEADER_SCAN_ROWS) + 1)
    ):
        return NdsRequestParseResult(matched=False, reason="header_not_found")

    for sheet in sheets:
        found = _sheet_looks_like_partner(sheet)
        if found is None:
            continue
        header_row, cols, form_kind = found
        return _parse_sheet(sheet, header_row, cols, form_kind)

    return NdsRequestParseResult(matched=False, reason="header_not_found")


class _PriceLine:
    __slots__ = ("supplier_inn", "amount")

    def __init__(self, supplier_inn: str, amount: Decimal) -> None:
        self.supplier_inn = supplier_inn
        self.amount = amount


def lines_for_pricing(application: ParsedApplication) -> list[Any]:
    return [_PriceLine(line.supplier_inn, line.amount) for line in application.lines]
