from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


MONTH_RU = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}


def director_name_key(name: str) -> str:
    text = " ".join(str(name).split()).casefold().replace("ё", "е")
    return text


def normalize_inn(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float):
        text = str(int(value)) if value == int(value) else str(value)
    else:
        text = str(value).strip()
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) < 8:
        return None
    return digits


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _as_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _as_money(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number == 0:
        return 0.0
    return number


def _header_map(row: tuple[Any, ...]) -> dict[str, int]:
    out: dict[str, int] = {}
    for idx, cell in enumerate(row):
        if cell is None:
            continue
        out[str(cell).strip()] = idx
    return out


def _cell(row: tuple[Any, ...], headers: dict[str, int], *names: str) -> Any:
    for name in names:
        idx = headers.get(name)
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def parse_period_header(label: str) -> str | None:
    text = " ".join(str(label).split()).casefold()
    parts = text.split()
    if len(parts) != 2:
        return None
    month = MONTH_RU.get(parts[0])
    if month is None:
        return None
    try:
        year = int(parts[1])
    except ValueError:
        return None
    return f"{year:04d}-{month:02d}"


def parse_svodnaya(content: bytes) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    shops: list[dict[str, Any]] = []
    shops.extend(_parse_company_sheet(wb, "Приоритетные компании", "priority"))
    shops.extend(_parse_company_sheet(wb, "Обслуживающие компании", "service"))
    shops.extend(_parse_new_shops(wb))
    payments = _parse_payments(wb, "Выплаты ПРИОРИТЕТНЫЕ")
    payments.extend(_parse_payments(wb, "Выплаты ОБСЛУЖИВАЮЩИЕ"))
    return {"shops": shops, "payments": payments}


def _parse_company_sheet(wb: Any, sheet_name: str, kind: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    if not rows:
        return []
    headers = _header_map(rows[0])
    items: list[dict[str, Any]] = []
    for row in rows[1:]:
        inn = normalize_inn(_cell(row, headers, "ИНН"))
        name = _as_text(_cell(row, headers, "Наименование компании"))
        if not inn or not name:
            continue
        items.append(
            {
                "inn": inn,
                "name": name,
                "kind": kind,
                "director_name": _as_text(_cell(row, headers, "ФИО директора")),
                "registered_at": _as_date(_cell(row, headers, "Дата регистрации")),
                "planned_payout": _as_money(_cell(row, headers, "Плановая сумма выплаты")),
                "company_status": _as_text(_cell(row, headers, "Статус компании")),
                "sale_priority": _as_text(_cell(row, headers, "Приоритет продажи")),
                "unreliable": _as_text(_cell(row, headers, "Недостоверка", "Недостоверность")),
                "ecsp_status": _as_text(_cell(row, headers, "ЭЦП")),
                "ecsp_until": _as_date(_cell(row, headers, "ЭЦП ДО")),
                "zsk": _as_text(_cell(row, headers, "ЗСК")),
                "banks": _as_text(_cell(row, headers, "Банки")),
                "accounts_status": _as_text(_cell(row, headers, "Статус счетов")),
                "manager": _as_text(_cell(row, headers, "Менеджер")),
                "phone": _as_text(_cell(row, headers, "Телефон")),
                "telegram": _as_text(_cell(row, headers, "Телеграм")),
                "source": "svodnaya",
            },
        )
    return items


def _parse_new_shops(wb: Any) -> list[dict[str, Any]]:
    if "НОВЫЕ ЛАВКИ" not in wb.sheetnames:
        return []
    rows = list(wb["НОВЫЕ ЛАВКИ"].iter_rows(values_only=True))
    if not rows:
        return []
    headers = _header_map(rows[0])
    items: list[dict[str, Any]] = []
    for row in rows[1:]:
        inn = normalize_inn(_cell(row, headers, "ИНН"))
        name = _as_text(_cell(row, headers, "НАИМЕНОВАНИЕ"))
        if not inn or not name:
            continue
        items.append(
            {
                "inn": inn,
                "name": name,
                "kind": "new",
                "director_name": _as_text(_cell(row, headers, "ДИР")),
                "in_touch": _as_text(_cell(row, headers, "НА СВЯЗИ?")),
                "ecsp_status": _as_text(_cell(row, headers, "ЭЦП")),
                "accounts_status": _as_text(_cell(row, headers, "СЧЕТА")),
                "banks": _as_text(_cell(row, headers, "БАНКИ")),
                "accountant": _as_text(_cell(row, headers, "Бухгалтер")),
                "treatment_status": _as_text(row[8] if len(row) > 8 else None),
                "passport": _as_text(_cell(row, headers, "Паспорт")),
                "inn_personal": normalize_inn(row[12] if len(row) > 12 else None),
                "snils": _as_text(_cell(row, headers, "СНИЛС")),
                "birth_date": _as_date(_cell(row, headers, "Дата д/р")),
                "source": "svodnaya",
            },
        )
    return items


def _parse_payments(wb: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    if not rows:
        return []
    header = rows[0]
    period_cols: list[tuple[int, str]] = []
    for idx, cell in enumerate(header):
        if cell is None:
            continue
        period = parse_period_header(str(cell))
        if period:
            period_cols.append((idx, period))
    items: list[dict[str, Any]] = []
    for row in rows[1:]:
        name = _as_text(row[1] if len(row) > 1 else None)
        director_name = _as_text(row[2] if len(row) > 2 else None)
        if not director_name:
            continue
        for idx, period in period_cols:
            if idx >= len(row):
                continue
            amount = _as_money(row[idx])
            if amount is None or amount <= 0:
                continue
            items.append(
                {
                    "shop_name": name,
                    "director_name": director_name,
                    "period_ym": period,
                    "amount": amount,
                },
            )
    return items
