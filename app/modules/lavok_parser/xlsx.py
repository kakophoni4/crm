from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.modules.leads.opt.contact_buyer import normalize_inn
from app.shared.exceptions import ValidationError

_SHEET_DATE_RE = re.compile(r"^(\d{2})\.(\d{2})\.(\d{4})$")

_HEADER_TO_FIELD: dict[str, str] = {
    "источник": "source",
    "название": "name",
    "инн": "inn",
    "цена": "price",
    "дата регистрации": "registered_at",
    "налог": "tax",
    "адрес и директор": "address_director",
    "суды": "courts",
    "долги / ил": "debts",
    "долги/ил": "debts",
    "достоверность егрюл": "egrul_reliability",
    "банкротство": "bankruptcy",
    "обороты": "turnover",
    "отчетность": "reporting",
    "лизинг / залоги": "leasing",
    "лизинг/залоги": "leasing",
    "зск": "zsk",
    "итог": "summary",
    "балл": "score",
    "первое появление": "first_seen",
    "продавец": "seller",
    "ссылка": "link",
    "companium": "companium",
    "статус егрюл": "egrul_status",
}

SNAPSHOT_FIELDS: tuple[str, ...] = (
    "source",
    "name",
    "price",
    "registered_at",
    "tax",
    "address_director",
    "courts",
    "debts",
    "egrul_reliability",
    "bankruptcy",
    "turnover",
    "reporting",
    "leasing",
    "zsk",
    "summary",
    "score",
    "first_seen",
    "seller",
    "link",
    "companium",
    "egrul_status",
)


@dataclass(frozen=True)
class ParsedLotRow:
    inn: str
    sheet_date: date
    fields: dict[str, str | None]


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower().replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    return text or None


def parse_sheet_date(title: str) -> date | None:
    match = _SHEET_DATE_RE.match((title or "").strip())
    if not match:
        return None
    day, month, year = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
    try:
        return date(year, month, day)
    except ValueError:
        return None


def parse_lavok_xlsx(content: bytes) -> list[ParsedLotRow]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValidationError(message="Не удалось прочитать Excel-файл парсера") from exc

    rows: list[ParsedLotRow] = []
    try:
        for worksheet in workbook.worksheets:
            sheet_date = parse_sheet_date(str(worksheet.title or ""))
            if sheet_date is None:
                continue
            header_map: dict[int, str] = {}
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row is None:
                continue
            for col_idx, value in enumerate(header_row, start=1):
                field = _HEADER_TO_FIELD.get(_norm_header(value))
                if field:
                    header_map[col_idx] = field
            if "inn" not in header_map.values():
                continue
            for excel_row in worksheet.iter_rows(min_row=2, values_only=True):
                raw: dict[str, str | None] = {}
                for col_idx, value in enumerate(excel_row, start=1):
                    field = header_map.get(col_idx)
                    if not field:
                        continue
                    raw[field] = _cell_text(value)
                inn = normalize_inn(raw.get("inn"))
                if not inn:
                    continue
                fields = {key: raw.get(key) for key in SNAPSHOT_FIELDS}
                rows.append(ParsedLotRow(inn=inn, sheet_date=sheet_date, fields=fields))
    finally:
        workbook.close()

    if not rows:
        raise ValidationError(message="В файле нет строк с датой листа ДД.ММ.ГГГГ и ИНН")
    return rows
