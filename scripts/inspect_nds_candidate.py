#!/usr/bin/env python3
"""Inspect why a storage spreadsheet would SKIP/HIT for NDS scan.

Usage:
  docker exec crm-staging-api python scripts/inspect_nds_candidate.py --name '2кв 2025г.xlsx'
  docker exec crm-staging-api python scripts/inspect_nds_candidate.py --key 'path/in/minio.xlsx'
"""

from __future__ import annotations

import argparse
import asyncio
import sys
from pathlib import Path

from sqlalchemy import text

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from openpyxl import load_workbook  # noqa: E402
from io import BytesIO  # noqa: E402

from app.modules.leads.opt.nds_request_parser import parse_nds_request_workbook  # noqa: E402
from app.shared.db import get_session_factory  # noqa: E402
from app.shared.storage import get_file_storage  # noqa: E402


async def _resolve_key(name: str | None, key: str | None) -> tuple[str, str]:
    if key:
        return key, key
    if not name:
        raise SystemExit("Need --name or --key")
    pat = name if ("%" in name or "_" in name) else f"%{name}%"
    session_factory = get_session_factory()
    async with session_factory() as session:
        row = (
            await session.execute(
                text(
                    """
                    SELECT storage_key, original_name, 'group_chat_files' AS src FROM group_chat_files
                    WHERE original_name ILIKE :n
                    UNION ALL
                    SELECT storage_key, original_name, 'uploaded_files' FROM uploaded_files
                    WHERE original_name ILIKE :n
                    LIMIT 5
                    """
                ),
                {"n": pat},
            )
        ).mappings().all()
    if not row:
        raise SystemExit(f"No file matching name={name!r}")
    print(f"Found {len(row)} match(es):")
    for r in row:
        print(f"  [{r['src']}] key={r['storage_key']} name={r['original_name']!r}")
    first = row[0]
    return str(first["storage_key"]), str(first["original_name"])


async def _amain(name: str | None, key: str | None) -> int:
    storage_key, display = await _resolve_key(name, key)
    storage = get_file_storage()
    content, ctype = await storage.get_bytes(storage_key)
    print(f"\nDownloaded {display!r} ({len(content)} bytes, ctype={ctype})")
    print(f"magic={content[:8]!r}")

    parsed = parse_nds_request_workbook(content)
    print(
        f"\nNDS parse: matched={parsed.matched} sheet={parsed.sheet_name!r} "
        f"reason={parsed.reason!r}",
    )
    if parsed.application:
        print(
            f"  buyer={parsed.application.buyer_inn} "
            f"lines={len(parsed.application.lines)}",
        )

    if content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        print("\nThis is legacy .xls (OLE) — openpyxl cannot read it.")
        return 0

    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        print(f"\nopenpyxl failed: {exc}")
        return 1

    print("\nFirst 15 rows of each sheet (for manual check):")
    for ws in wb.worksheets:
        print(f"--- sheet {ws.title!r} max_row={ws.max_row} max_col={ws.max_column}")
        for r in range(1, min(16, (ws.max_row or 0) + 1)):
            vals = []
            for c in range(1, min(15, (ws.max_column or 0) + 1)):
                v = ws.cell(r, c).value
                if v is not None:
                    vals.append(f"{c}:{v!r}")
            print(f"  row{r}: {', '.join(vals) if vals else '(empty)'}")
    wb.close()
    return 0


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument(
        "--name",
        help="Substring or ILIKE pattern for original_name (auto %%wrap%%)",
    )
    p.add_argument("--key", help="storage_key")
    args = p.parse_args()
    raise SystemExit(asyncio.run(_amain(args.name, args.key)))


if __name__ == "__main__":
    main()
