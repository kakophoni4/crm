#!/usr/bin/env python3
"""Audit Excel header rows in CRM storage (partner focus, no OPT totals).

For every spreadsheet candidate prints/classifies the first header-like row
so we can see what partner-like files are being skipped and why.

Kinds:
  nds_request   — «Заявка на НДС» / ЗАПРОС НДС (ИНН покупателя + стоимость покупки + ИНН продавца)
  partner_forma — Forma_zayavki VAT form (ИНН организации + сумма НДС + …)
  opt_zayavka   — CRM OPT upload (ИНН компании-продавца / сумма покупок) — marked only, not summed
  xls_legacy    — old .xls
  other         — something else

Usage:
  docker cp scripts/audit_storage_headers.py crm-staging-api:/app/scripts/
  docker exec -e PYTHONUNBUFFERED=1 crm-staging-api \
    python scripts/audit_storage_headers.py > /tmp/header_audit.txt 2>&1
  grep -A200 '^=== BY KIND' /tmp/header_audit.txt
  grep -A200 '^=== OTHER HEADER' /tmp/header_audit.txt
"""

from __future__ import annotations

import argparse
import asyncio
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path

from sqlalchemy import text

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from openpyxl import load_workbook  # noqa: E402

from app.shared.db import get_session_factory  # noqa: E402
from app.shared.storage import get_file_storage  # noqa: E402


def _log(msg: str) -> None:
    print(msg, flush=True)


def _norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


@dataclass
class HeaderSample:
    kind: str
    signature: str
    sheet: str
    row_idx: int
    examples: list[str] = field(default_factory=list)
    count: int = 0


def _classify(headers: list[str]) -> str:
    blob = " | ".join(headers)
    norms = set(headers)

    if "инн покупателя" in blob and "стоимость покупки" in blob and "инн продавца" in blob:
        return "nds_request"
    if "инн покупателя" in blob and ("инн продавца" in blob or "инн организации" in blob):
        if "сумма ндс" in blob or "сумма (в т.ч. ндс)" in blob or "стоимость покупки" in blob:
            return "nds_or_partner_near"

    partner_hits = 0
    if "сумма ндс" in norms:
        partner_hits += 1
    if "наименование покупателя" in norms or "инн покупателя" in norms:
        partner_hits += 1
    if "инн организации" in norms or "наименование организации" in norms:
        partner_hits += 1
    if "сумма (в т.ч. ндс)" in norms or "сумма в т.ч. ндс" in norms:
        partner_hits += 1
    if partner_hits >= 2:
        return "partner_forma"

    if "инн компании-продавца" in blob or (
        "сумма покупок" in blob and "инн" in blob and "продав" in blob
    ):
        return "opt_zayavka"
    if "№ документа" in blob or "сумма без ндс" in norms:
        return "crm_registry"
    if "раздел" in blob or "наименование товара" in blob:
        return "other_business"
    return "other"


def _extract_headers(content: bytes) -> tuple[str, str, int, list[str]] | tuple[str, None, None, None]:
    if content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls_legacy", None, None, None
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        return f"open_fail:{type(exc).__name__}", None, None, None

    try:
        best: tuple[int, str, int, list[str]] | None = None
        for ws in wb.worksheets:
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=8, max_col=20, values_only=True),
                start=1,
            ):
                cells = [_norm(v) for v in row if v is not None and str(v).strip()]
                if len(cells) < 3:
                    continue
                # Prefer rows that look like headers (contain инн/сумма/дата/наименование)
                score = sum(
                    1
                    for c in cells
                    if any(k in c for k in ("инн", "сумм", "стоим", "дата", "наимен", "продав", "покуп"))
                )
                if score < 2:
                    continue
                kind = _classify(cells)
                # Prefer classified non-other
                rank = 0 if kind.startswith(("nds", "partner", "opt")) else 1
                cand = (rank, ws.title, row_idx, cells)
                if best is None or cand[0] < best[0] or (
                    cand[0] == best[0] and score > 2 and cand[2] < best[2]
                ):
                    best = cand
        if best is None:
            return "no_header_row", None, None, None
        _rank, sheet, row_idx, cells = best
        return _classify(cells), sheet, row_idx, cells
    finally:
        wb.close()


async def _candidates(*, limit: int | None) -> list[tuple[str, str, str]]:
    session_factory = get_session_factory()
    found: dict[str, tuple[str, str, str]] = {}
    async with session_factory() as session:
        for sql in (
            """
            SELECT storage_key, original_name, 'uploaded_files'
            FROM uploaded_files
            WHERE lower(original_name) LIKE '%.xlsx'
               OR lower(original_name) LIKE '%.xlsm'
               OR lower(original_name) LIKE '%.xls'
            """,
            """
            SELECT storage_key, original_name, 'group_chat_files'
            FROM group_chat_files
            WHERE lower(original_name) LIKE '%.xlsx'
               OR lower(original_name) LIKE '%.xlsm'
               OR lower(original_name) LIKE '%.xls'
            """,
            """
            SELECT att->>'storage_key',
                   COALESCE(att->>'filename', att->>'name', 'attachment.xlsx'),
                   'message_attachments'
            FROM messages m
            CROSS JOIN LATERAL jsonb_array_elements(
              CASE WHEN jsonb_typeof(m.attachments)='array' THEN m.attachments ELSE '[]'::jsonb END
            ) att
            WHERE coalesce(att->>'storage_key','') <> ''
              AND (
                lower(COALESCE(att->>'filename', att->>'name', '')) LIKE '%.xlsx'
                OR lower(COALESCE(att->>'filename', att->>'name', '')) LIKE '%.xlsm'
                OR lower(COALESCE(att->>'filename', att->>'name', '')) LIKE '%.xls'
              )
            """,
        ):
            for key, name, src in (await session.execute(text(sql))).all():
                k = str(key or "").strip()
                if k and k not in found:
                    found[k] = (k, str(name or k), str(src))
    items = sorted(found.values(), key=lambda x: x[1].lower())
    if limit is not None:
        items = items[:limit]
    return items


async def _amain(*, limit: int | None, name_filter: str | None) -> int:
    cands = await _candidates(limit=limit)
    if name_filter:
        needle = name_filter.lower()
        cands = [c for c in cands if needle in c[1].lower()]
    storage = get_file_storage()
    total = len(cands)
    _log(f"Candidates: {total}")

    by_kind: dict[str, int] = defaultdict(int)
    # signature -> sample
    groups: dict[tuple[str, str], HeaderSample] = {}

    for i, (key, name, src) in enumerate(cands, start=1):
        pct = (100 * i) // total if total else 100
        try:
            content, _ = await storage.get_bytes(key)
        except Exception as exc:  # noqa: BLE001
            by_kind["download_fail"] += 1
            _log(f"[{i}/{total} {pct}%] DOWNLOAD_FAIL {name!r}: {exc}")
            continue

        kind, sheet, row_idx, cells = _extract_headers(content)
        by_kind[kind] += 1
        if cells is None:
            _log(f"[{i}/{total} {pct}%] [{src}] {name!r} → {kind}")
            continue

        sig = " | ".join(cells[:12])
        gkey = (kind, sig)
        sample = groups.get(gkey)
        if sample is None:
            sample = HeaderSample(
                kind=kind,
                signature=sig,
                sheet=sheet or "?",
                row_idx=row_idx or 0,
            )
            groups[gkey] = sample
        sample.count += 1
        if len(sample.examples) < 5:
            sample.examples.append(name)

        # Compact line for partner-ish / interesting
        if kind in {"nds_request", "partner_forma", "nds_or_partner_near", "other"} or any(
            x in name.lower() for x in ("заявк", "запрос", "форма")
        ):
            _log(
                f"[{i}/{total} {pct}%] [{src}] {name!r} → {kind} "
                f"sheet={sheet!r} row={row_idx} :: {sig[:160]}",
            )
        elif i % 25 == 0:
            _log(f"[{i}/{total} {pct}%] … {kind}")

    _log("")
    _log("=== BY KIND ===")
    for kind, n in sorted(by_kind.items(), key=lambda kv: (-kv[1], kv[0])):
        _log(f"  {kind}: {n}")

    _log("")
    _log("=== NDS / PARTNER / NEAR signatures ===")
    for sample in sorted(
        (s for s in groups.values() if s.kind in {"nds_request", "partner_forma", "nds_or_partner_near"}),
        key=lambda s: (-s.count, s.kind),
    ):
        _log(f"\n[{sample.kind}] x{sample.count} sheet={sample.sheet!r} row={sample.row_idx}")
        _log(f"  headers: {sample.signature}")
        for ex in sample.examples:
            _log(f"    - {ex}")

    _log("")
    _log("=== OTHER HEADER groups (top 40 by count) ===")
    others = sorted(
        (s for s in groups.values() if s.kind in {"other", "other_business"}),
        key=lambda s: -s.count,
    )[:40]
    for sample in others:
        _log(f"\n[{sample.kind}] x{sample.count} sheet={sample.sheet!r} row={sample.row_idx}")
        _log(f"  headers: {sample.signature}")
        for ex in sample.examples:
            _log(f"    - {ex}")

    _log("")
    _log("Done. OPT is listed as opt_zayavka for visibility only — not totaled here.")
    return 0


def main() -> None:
    try:
        sys.stdout.reconfigure(line_buffering=True)  # type: ignore[attr-defined]
    except Exception:
        pass
    p = argparse.ArgumentParser(description="Audit spreadsheet headers in storage")
    p.add_argument("--limit", type=int, default=None)
    p.add_argument("--name-filter", help="Only names containing this substring")
    args = p.parse_args()
    raise SystemExit(asyncio.run(_amain(limit=args.limit, name_filter=args.name_filter)))


if __name__ == "__main__":
    main()
