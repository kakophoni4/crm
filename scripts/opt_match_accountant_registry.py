#!/usr/bin/env python3
"""Match accountant-filled ОБЩИЙ ТАВРИДА.xlsx to CRM OPT lines.

RULE: every document number in the Excel is already filed — keep 100%.
Only CRM lines that are NOT in the Excel may be sent to 1C as a NEW remainder order.

Usage on VPS (copy xlsx into container or mount):
  docker cp "ОБЩИЙ ТАВРИДА.xlsx" crm-staging-api:/tmp/tavrida.xlsx
  docker cp scripts/opt_match_accountant_registry.py crm-staging-api:/app/scripts/

  # dry compare
  docker exec crm-staging-api python /app/scripts/opt_match_accountant_registry.py \\
    --xlsx /tmp/tavrida.xlsx --order-ids 178,179,253

  # write document_number from Excel onto matched CRM lines
  docker exec crm-staging-api python /app/scripts/opt_match_accountant_registry.py \\
    --xlsx /tmp/tavrida.xlsx --order-ids 178,179,253 --apply-docs

  # also scan all active Таврида orders on lead 363
  docker exec crm-staging-api python /app/scripts/opt_match_accountant_registry.py \\
    --xlsx /tmp/tavrida.xlsx --lead-id 363 --buyer-inn 9102191311 --apply-docs
"""

from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from openpyxl import load_workbook

DOC_RE = re.compile(r"^([A-Za-zА-Яа-яЁё]{1,4}-\d{5,}|\d{2}-\d{5,}|[A-ZА-Я]{2}\d{10,})$")


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, date):
        return v.strftime("%d.%m.%Y")
    return str(v).strip()


def _parse_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    text = str(v).strip()
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v)).quantize(Decimal("0.01"))
    text = (
        str(v)
        .replace("\xa0", "")
        .replace(" ", "")
        .replace("₽", "")
        .replace(",", ".")
    )
    text = re.sub(r"[^0-9.\-]", "", text)
    if not text or text in {".", "-", "-."}:
        return None
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except Exception:
        return None


def _normalize_inn(v: Any) -> str | None:
    if v is None:
        return None
    digits = re.sub(r"\D", "", str(v))
    if len(digits) in {10, 12}:
        return digits
    return None


def _find_doc(v: Any) -> str | None:
    text = _cell_str(v).upper().replace("Ё", "Е")
    if not text:
        return None
    if DOC_RE.match(text):
        return text
    m = re.search(r"([A-ZА-Я]{1,4}-\d{5,}|\d{2}-\d{5,}|[A-ZА-Я]{2}\d{10,})", text)
    return m.group(1) if m else None


@dataclass
class AccRow:
    sheet: str
    row: int
    buyer_inn: str | None
    supplier_inn: str | None
    document_date: date | None
    amount: Decimal | None
    document_number: str | None
    source: str


def parse_accountant_xlsx(path: Path) -> list[AccRow]:
    """Parse both Mole-export blocks and filled application-form blocks."""
    wb = load_workbook(path, data_only=True)
    out: list[AccRow] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = int(ws.max_row or 0)
        max_col = min(int(ws.max_column or 0), 20)

        for r in range(1, max_row + 1):
            # Format A: № документа | дата | покупатель | ИНН | … | поставщик | ИНН | … | сумма
            doc_a = _find_doc(ws.cell(r, 1).value)
            date_b = _parse_date(ws.cell(r, 2).value)
            buyer_d = _normalize_inn(ws.cell(r, 4).value)
            supplier_g = _normalize_inn(ws.cell(r, 7).value)
            amount_i = _parse_amount(ws.cell(r, 9).value)

            if doc_a and date_b and amount_i and amount_i > 0 and (supplier_g or buyer_d):
                out.append(
                    AccRow(
                        sheet=sheet_name,
                        row=r,
                        buyer_inn=buyer_d,
                        supplier_inn=supplier_g,
                        document_date=date_b,
                        amount=amount_i,
                        document_number=doc_a,
                        source="registry_export",
                    )
                )
                continue

            # Format B: form with date in F, amount G, supplier INN in I, SF no in M (or K)
            buyer_c = _normalize_inn(ws.cell(r, 3).value)
            date_f = _parse_date(ws.cell(r, 6).value)
            amount_g = _parse_amount(ws.cell(r, 7).value)
            supplier_i = _normalize_inn(ws.cell(r, 9).value)
            doc_m = _find_doc(ws.cell(r, 13).value) or _find_doc(ws.cell(r, 11).value)

            if buyer_c and date_f and amount_g and amount_g > 0 and supplier_i and doc_m:
                out.append(
                    AccRow(
                        sheet=sheet_name,
                        row=r,
                        buyer_inn=buyer_c,
                        supplier_inn=supplier_i,
                        document_date=date_f,
                        amount=amount_g,
                        document_number=doc_m,
                        source="filled_form",
                    )
                )

    return out


def _key(supplier: str | None, d: date | None, amount: Decimal | None) -> str:
    return f"{supplier or '?'}|{d.isoformat() if d else '?'}|{amount if amount is not None else '?'}"


async def run(
    *,
    xlsx: Path,
    order_ids: list[int],
    lead_id: int | None,
    buyer_inn: str | None,
    apply_docs: bool,
) -> dict[str, Any]:
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload

    from app.modules.db.models.lead_opt_order import LeadOptOrder
    from app.shared.db import get_session_factory

    acc_rows = parse_accountant_xlsx(xlsx)
    by_key: dict[str, list[AccRow]] = {}
    for ar in acc_rows:
        by_key.setdefault(_key(ar.supplier_inn, ar.document_date, ar.amount), []).append(ar)

    sf = get_session_factory()
    report: dict[str, Any] = {
        "xlsx": str(xlsx),
        "accountant_rows": len(acc_rows),
        "accountant_with_doc": sum(1 for a in acc_rows if a.document_number),
        "accountant_sum": str(sum((a.amount or Decimal(0)) for a in acc_rows)),
        "orders": [],
        "send_only_lines": [],
        "applied": False,
    }

    async with sf() as session:
        q = select(LeadOptOrder).options(selectinload(LeadOptOrder.lines)).where(
            LeadOptOrder.deleted_at.is_(None)
        )
        if order_ids:
            q = q.where(LeadOptOrder.id.in_(order_ids))
        if lead_id is not None:
            q = q.where(LeadOptOrder.lead_id == lead_id)
        if buyer_inn:
            q = q.where(LeadOptOrder.buyer_inn == buyer_inn)

        orders = (await session.execute(q.order_by(LeadOptOrder.order_no, LeadOptOrder.id))).scalars().all()
        used: set[int] = set()  # id(AccRow) via row identity: (sheet,row)

        for order in orders:
            matched = []
            missing = []
            for line in sorted(order.lines, key=lambda x: x.line_no):
                amt = Decimal(str(line.amount)).quantize(Decimal("0.01")) if line.amount is not None else None
                k = _key(line.supplier_inn, line.document_date, amt)
                pick: AccRow | None = None
                for cand in by_key.get(k, []):
                    uid = (cand.sheet, cand.row)
                    if uid in used:
                        continue
                    pick = cand
                    used.add(uid)
                    break

                item = {
                    "line_no": line.line_no,
                    "line_id": line.id,
                    "crm_id": line.crm_id,
                    "supplier_inn": line.supplier_inn,
                    "date": line.document_date.isoformat() if line.document_date else None,
                    "amount": float(amt) if amt is not None else None,
                    "crm_doc_before": line.document_number,
                }
                if pick is None:
                    item["status"] = "SEND_NEW"
                    missing.append(item)
                    report["send_only_lines"].append(
                        {
                            "order_id": order.id,
                            "order_no": order.order_no,
                            **item,
                        }
                    )
                else:
                    item["status"] = "KEEP"
                    item["acc_doc"] = pick.document_number
                    item["acc_row"] = pick.row
                    item["acc_source"] = pick.source
                    if apply_docs and pick.document_number:
                        line.document_number = pick.document_number
                    matched.append(item)

            report["orders"].append(
                {
                    "id": order.id,
                    "order_no": order.order_no,
                    "crm_id": order.crm_id,
                    "buyer_inn": order.buyer_inn,
                    "file": order.source_filename,
                    "volume": float(order.total_volume or 0),
                    "keep": len(matched),
                    "send_new": len(missing),
                    "matched": matched,
                    "missing": missing,
                }
            )

        if apply_docs:
            await session.commit()
            report["applied"] = True

    # Accountant rows not used by any selected CRM line
    orphans = []
    for ar in acc_rows:
        if (ar.sheet, ar.row) not in used:
            orphans.append(
                {
                    "sheet": ar.sheet,
                    "row": ar.row,
                    "buyer_inn": ar.buyer_inn,
                    "supplier_inn": ar.supplier_inn,
                    "date": ar.document_date.isoformat() if ar.document_date else None,
                    "amount": float(ar.amount) if ar.amount is not None else None,
                    "doc": ar.document_number,
                    "source": ar.source,
                }
            )
    report["accountant_not_matched_to_selected_orders"] = orphans
    return report


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--order-ids", default="")
    parser.add_argument("--lead-id", type=int, default=None)
    parser.add_argument("--buyer-inn", default=None)
    parser.add_argument("--apply-docs", action="store_true")
    parser.add_argument("--out", default="")
    parser.add_argument("--no-db", action="store_true")
    args = parser.parse_args()

    path = Path(args.xlsx)
    if not path.is_file():
        print(f"FILE NOT FOUND: {path}")
        return 1

    if args.no_db:
        rows = parse_accountant_xlsx(path)
        print(f"rows={len(rows)} sum={sum((r.amount or 0) for r in rows)}")
        for r in rows:
            print(
                f"  r{r.row} {r.document_number} {r.document_date} "
                f"buyer={r.buyer_inn} supplier={r.supplier_inn} amount={r.amount} [{r.source}]"
            )
        return 0

    order_ids = [int(x) for x in args.order_ids.split(",") if x.strip()]
    if not order_ids and args.lead_id is None:
        order_ids = [178, 179, 253]

    report = asyncio.run(
        run(
            xlsx=path,
            order_ids=order_ids,
            lead_id=args.lead_id,
            buyer_inn=args.buyer_inn,
            apply_docs=args.apply_docs,
        )
    )

    print(f"xlsx={report['xlsx']}")
    print(
        f"accountant_rows={report['accountant_rows']} "
        f"with_doc={report['accountant_with_doc']} sum={report['accountant_sum']}"
    )
    print(f"applied={report['applied']}")

    for o in report["orders"]:
        print(
            f"\norder={o['id']} no={o['order_no']} buyer={o['buyer_inn']} "
            f"vol={o['volume']} KEEP={o['keep']} SEND_NEW={o['send_new']} file={o['file']}"
        )
        for m in o["matched"]:
            print(
                f"  KEEP  L{m['line_no']} {m['supplier_inn']} {m['date']} {m['amount']} "
                f"-> {m.get('acc_doc')} (was {m.get('crm_doc_before')})"
            )
        for m in o["missing"]:
            print(
                f"  SEND  L{m['line_no']} {m['supplier_inn']} {m['date']} {m['amount']} "
                f"(crm_doc={m.get('crm_doc_before')})"
            )

    send = report["send_only_lines"]
    print(f"\n=== SEND ONLY ({len(send)} lines) — create NEW remainder order(s), do NOT remake KEEP orders ===")
    for m in send:
        print(
            f"  from order {m['order_no']}/{m['order_id']} L{m['line_no']} "
            f"{m['supplier_inn']} {m['date']} {m['amount']}"
        )

    orphans = report["accountant_not_matched_to_selected_orders"]
    print(f"\n=== accountant rows not in selected CRM orders ({len(orphans)}) ===")
    for o in orphans[:50]:
        print(f"  r{o['row']} {o['doc']} {o['date']} {o['supplier_inn']} {o['amount']} [{o['source']}]")
    if len(orphans) > 50:
        print(f"  ... +{len(orphans) - 50} more")

    out = Path(args.out) if args.out else path.with_suffix(".match.json")
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"\nWrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
