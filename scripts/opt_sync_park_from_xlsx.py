#!/usr/bin/env python3
"""Sync OPT lavka categories/rates from «Парк компаний» spreadsheet.

Reads columns:
  A — category label (Абсолют / Оптима / Техничка)
  B — company name
  C — INN
  K — commission rate (fraction, e.g. 0.035 = 3.5%)

Companies listed in the park file get category_code + commission_rate_percent + is_active.
With --deactivate-missing (default), all other opt_units are soft-deactivated so OPT
orders cannot resolve those INNs (get_unit_by_inn filters is_active).

Usage:
  py scripts/opt_sync_park_from_xlsx.py
  py scripts/opt_sync_park_from_xlsx.py --xlsx "Парк_компаний_на_2КВ2026 13.07.xlsx"
  py scripts/opt_sync_park_from_xlsx.py --sql scripts/deploy/seed-opt-park-categories.sql
  py scripts/opt_sync_park_from_xlsx.py --keep-missing   # do not deactivate others
"""

from __future__ import annotations

import argparse
import json
import re
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

PARK_CATEGORY_LABEL_TO_CODE: dict[str, str] = {
    "абсолют": "L",
    "оптима": "O",
    "техничка": "TECH",
}

_INN_RE = re.compile(r"^\d{10}(\d{2})?$")
_ROOT = Path(__file__).resolve().parents[1]
_DEFAULT_JSON = _ROOT / "scripts" / "opt_park_categories.json"


def _find_default_xlsx() -> Path:
    preferred = sorted(
        _ROOT.glob("Парк_компаний*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if preferred:
        return preferred[0]
    legacy = _ROOT / "Парк_компаний_на_2КВ2026 09.07.xlsx"
    return legacy


def _normalize_inn(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float):
        text = str(int(value))
    elif isinstance(value, int):
        text = str(value)
    else:
        text = str(value).strip().replace(" ", "")
        if text.endswith(".0"):
            text = text[:-2]
    return text if _INN_RE.match(text) else None


def _normalize_rate(value: object) -> Decimal | None:
    if value is None:
        return None
    try:
        rate = Decimal(str(value))
    except Exception:
        return None
    if rate <= 0:
        return None
    if rate < 1:
        rate = (rate * Decimal("100")).quantize(Decimal("0.01"))
    return rate.quantize(Decimal("0.01"))


def _map_category(label: object) -> tuple[str, str] | None:
    text = str(label or "").strip()
    if not text:
        return None
    code = PARK_CATEGORY_LABEL_TO_CODE.get(text.casefold())
    if code is None:
        raise ValueError(f"Unknown park category label: {text!r}")
    return text, code


def load_park_entries(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    entries: list[dict[str, object]] = []
    seen: set[str] = set()
    for row_idx in range(2, (ws.max_row or 0) + 1):
        category_raw = ws.cell(row_idx, 1).value
        name_raw = ws.cell(row_idx, 2).value
        inn = _normalize_inn(ws.cell(row_idx, 3).value)
        rate = _normalize_rate(ws.cell(row_idx, 11).value)
        if not inn:
            continue
        mapped = _map_category(category_raw)
        if mapped is None or not name_raw:
            continue
        category_label, category_code = mapped
        if inn in seen:
            continue
        seen.add(inn)
        entries.append(
            {
                "inn": inn,
                "name": str(name_raw).strip(),
                "category_label": category_label,
                "category_code": category_code,
                "commission_rate_percent": float(rate) if rate is not None else None,
            },
        )
    wb.close()
    if not entries:
        raise SystemExit(f"No park rows found in {path}")
    return entries


def write_sql(
    entries: list[dict[str, object]],
    path: Path,
    *,
    deactivate_missing: bool,
) -> None:
    lines = [
        "-- Categories/rates from «Парк компаний» spreadsheet",
        "-- Run after seed-opt-lavki.sql on server",
        "",
        "BEGIN;",
        "",
    ]
    park_inns = [str(entry["inn"]) for entry in entries]
    for entry in entries:
        inn = str(entry["inn"])
        name = str(entry["name"]).replace("'", "''")
        category_code = str(entry["category_code"])
        rate = entry.get("commission_rate_percent")
        rate_sql = "NULL" if rate is None else str(rate)
        lines.append(
            f"-- {entry['category_label']}: {name}\n"
            "UPDATE opt_units\n"
            f"SET name = '{name}',\n"
            f"    category_code = '{category_code}',\n"
            f"    commission_rate_percent = {rate_sql},\n"
            "    is_active = TRUE\n"
            f"WHERE inn = '{inn}';\n"
            f"INSERT INTO opt_units (inn, name, category_code, commission_rate_percent, is_active)\n"
            f"SELECT '{inn}', '{name}', '{category_code}', {rate_sql}, TRUE\n"
            f"WHERE NOT EXISTS (SELECT 1 FROM opt_units WHERE inn = '{inn}');\n"
        )

    if deactivate_missing and park_inns:
        inn_list = ",\n  ".join(f"'{inn}'" for inn in park_inns)
        lines.append(
            "-- Soft-deactivate lavki missing from park (blocks OPT order resolution)\n"
            "UPDATE opt_units\n"
            "SET is_active = FALSE\n"
            "WHERE is_active IS TRUE\n"
            f"  AND inn NOT IN (\n  {inn_list}\n);\n"
        )

    lines.extend(["", "COMMIT;", ""])
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync park categories from Excel")
    parser.add_argument("--xlsx", type=Path, default=None)
    parser.add_argument("--json", type=Path, default=_DEFAULT_JSON)
    parser.add_argument("--sql", type=Path, default=None)
    parser.add_argument(
        "--deactivate-missing",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Soft-deactivate opt_units whose INN is not in the park file (default: on)",
    )
    args = parser.parse_args()

    xlsx = args.xlsx or _find_default_xlsx()
    if not xlsx.is_file():
        raise SystemExit(f"File not found: {xlsx}")

    entries = load_park_entries(xlsx)
    args.json.write_text(
        json.dumps(entries, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Read {len(entries)} park companies from {xlsx.name}")
    print(f"Wrote {args.json}")

    sql_path = args.sql or (_ROOT / "scripts" / "deploy" / "seed-opt-park-categories.sql")
    write_sql(entries, sql_path, deactivate_missing=args.deactivate_missing)
    print(f"Wrote {sql_path}")
    if args.deactivate_missing:
        print("SQL includes soft-deactivation of opt_units not present in the park file.")


if __name__ == "__main__":
    main()
