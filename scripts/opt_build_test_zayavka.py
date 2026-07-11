#!/usr/bin/env python3
"""Build OPT test xlsx from synthetic spec (no real company data in repo).

Usage:
  py scripts/opt_build_test_zayavka.py

Output (gitignored):
  scripts/fixtures/opt-test-crm.xlsx
"""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

_ROOT = Path(__file__).resolve().parents[1]
_DEFAULT_SPEC = _ROOT / "scripts" / "fixtures" / "opt-test-zayavka-spec.json"
_DEFAULT_OUT = _ROOT / "scripts" / "fixtures" / "opt-test-crm.xlsx"


def _parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def build_workbook_from_spec(spec: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = str(spec.get("sheet_title") or "Sheet1")

    row2 = spec.get("header_row2") or {}
    row3 = spec.get("header_row3") or {}
    for col_key, value in row2.items():
        ws[f"{col_key}2"] = value
    for col_key, value in row3.items():
        ws[f"{col_key}3"] = value

    buyer_inn = str(spec["buyer"]["inn"])
    row = 4
    for line in spec["lines"]:
        ws.cell(row, 2, str(line["supplier_inn"]))
        ws.cell(row, 3, buyer_inn)
        ws.cell(row, 4, _parse_date(str(line["document_date"])))
        amount = line["amount"]
        if isinstance(amount, str):
            amount = float(Decimal(amount))
        ws.cell(row, 5, float(amount))
        row += 1

    total_row = int(spec.get("total_row") or row + 4)
    total_amount = spec.get("total_amount")
    if total_amount is not None:
        ws.cell(total_row, 5, float(total_amount))

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="Build synthetic OPT test xlsx")
    parser.add_argument("--spec", type=Path, default=_DEFAULT_SPEC)
    parser.add_argument("--out", type=Path, default=_DEFAULT_OUT)
    args = parser.parse_args()

    args.out.parent.mkdir(parents=True, exist_ok=True)
    spec = json.loads(args.spec.read_text(encoding="utf-8"))
    wb = build_workbook_from_spec(spec)
    wb.save(args.out)
    print(f"Built from spec {args.spec} -> {args.out}")

    import sys

    sys.path.insert(0, str(_ROOT))
    from app.modules.leads.opt.parser import parse_application_workbook

    parsed = parse_application_workbook(args.out.read_bytes())
    print(f"Parsed: buyer INN {parsed.buyer_inn}, lines {len(parsed.lines)}")
    for idx, line in enumerate(parsed.lines, start=1):
        print(f"  {idx}. {line.supplier_inn}  {line.document_date}  {line.amount}")


if __name__ == "__main__":
    main()
