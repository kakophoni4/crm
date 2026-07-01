#!/usr/bin/env python3
"""Sync opt_units seed JSON (and optional SQL) from «лавки Ване.xlsx».

Usage:
  py scripts/opt_sync_lavki_from_xlsx.py
  py scripts/opt_sync_lavki_from_xlsx.py --xlsx "лавки Ване.xlsx" --sql scripts/deploy/seed-opt-lavki.sql
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

_INN_RE = re.compile(r"^\d{10}(\d{2})?$")
_ROOT = Path(__file__).resolve().parents[1]
_DEFAULT_JSON = _ROOT / "scripts" / "opt_units_vane.json"


def _normalize_inn(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float):
        text = str(int(value))
    elif isinstance(value, int):
        text = str(value)
    else:
        text = str(value).strip().replace(" ", "")
        if text.endswith(".0"):
            text = text[:-2]
    return text if _INN_RE.match(text) else None


def _find_lavki_xlsx(explicit: Path | None) -> Path:
    if explicit is not None:
        if not explicit.is_file():
            raise SystemExit(f"File not found: {explicit}")
        return explicit
    candidates: list[Path] = []
    for path in _ROOT.glob("*.xlsx"):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            wb.close()
        except Exception:
            continue
        if cols == 2 and rows >= 50:
            candidates.append(path)
    if not candidates:
        raise SystemExit("Could not find lavki xlsx (2 columns, 50+ rows) in project root")
    if len(candidates) == 1:
        return candidates[0]
    for path in candidates:
        if "ване" in path.name.lower() or "лавки" in path.name.lower():
            return path
    return candidates[0]


def load_units_from_xlsx(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    units: list[dict[str, str]] = []
    seen: set[str] = set()
    skipped = 0
    for row_idx in range(1, (ws.max_row or 0) + 1):
        name_raw = ws.cell(row_idx, 1).value
        inn = _normalize_inn(ws.cell(row_idx, 2).value)
        if not inn or not name_raw:
            skipped += 1
            continue
        name = str(name_raw).strip()
        if not name or inn in seen:
            continue
        seen.add(inn)
        units.append({"name": name, "inn": inn})
    wb.close()
    if not units:
        raise SystemExit(f"No valid lavki rows in {path} (skipped {skipped})")
    return units


def write_sql(units: list[dict[str, str]], path: Path) -> None:
    lines = [
        "-- Upsert lavki from opt_units_vane.json / лавки Ване.xlsx",
        "-- Run on server:",
        "--   docker exec -i crm-staging-postgres psql -U crm -d crm < scripts/deploy/seed-opt-lavki.sql",
        "",
    ]
    for unit in units:
        inn = unit["inn"]
        name = unit["name"].replace("'", "''")
        lines.append(
            "INSERT INTO opt_units (inn, name, category_code, is_active)\n"
            f"VALUES ('{inn}', '{name}', 'TECH', TRUE)\n"
            "ON CONFLICT (inn) DO UPDATE SET\n"
            "  name = EXCLUDED.name,\n"
            "  category_code = COALESCE(opt_units.category_code, 'TECH'),\n"
            "  is_active = TRUE;"
        )
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync lavki JSON/SQL from Excel")
    parser.add_argument("--xlsx", type=Path, default=None, help="Path to лавки xlsx")
    parser.add_argument("--json", type=Path, default=_DEFAULT_JSON, help="Output JSON path")
    parser.add_argument("--sql", type=Path, default=None, help="Optional SQL upsert file")
    args = parser.parse_args()

    xlsx_path = _find_lavki_xlsx(args.xlsx)
    units = load_units_from_xlsx(xlsx_path)
    args.json.write_text(json.dumps(units, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Read {len(units)} lavki from {xlsx_path.name}")
    print(f"Wrote {args.json}")

    if args.sql is not None:
        write_sql(units, args.sql)
        print(f"Wrote {args.sql}")


if __name__ == "__main__":
    main()
