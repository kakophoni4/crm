#!/usr/bin/env python3
"""Find OPT orders truncated by the old Excel parser (mid-file empty rows).

Re-reads source xlsx (from chat attachment / storage) with the FIXED parser and
compares line count / volume to what is stored on the order.

Usage on VPS:
  docker exec -e PYTHONUNBUFFERED=1 crm-staging-api \\
    python scripts/opt_scan_truncated_orders.py

  docker exec -e PYTHONUNBUFFERED=1 crm-staging-api \\
    python scripts/opt_scan_truncated_orders.py --lead-id 363

  docker exec -e PYTHONUNBUFFERED=1 crm-staging-api \\
    python scripts/opt_scan_truncated_orders.py --only-suspects --limit 500

  # Also scan chat xlsx that would truncate under the OLD parser (no DB order needed):
  docker exec -e PYTHONUNBUFFERED=1 crm-staging-api \\
    python scripts/opt_scan_truncated_orders.py --scan-files --limit 200
"""

from __future__ import annotations

import argparse
import asyncio
import sys
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from app.modules.leads.opt.contact_buyer import (  # noqa: E402
    normalize_inn,
    parse_decimal,
    parse_excel_date,
)
from app.modules.leads.opt.parser import parse_application_workbook  # noqa: E402
from app.shared.db import get_session_factory  # noqa: E402
from app.shared.storage import get_file_storage  # noqa: E402


def _log(msg: str) -> None:
    print(msg, flush=True)


def _money(value: Decimal) -> str:
    return f"{value.quantize(Decimal('0.01')):,.2f}".replace(",", " ")


def parse_legacy_truncated(content: bytes) -> list[tuple[str, str, date, Decimal]]:
    """Old behaviour: stop at first empty/incomplete row after data started."""
    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active
    lines: list[tuple[str, str, date, Decimal]] = []
    buyer_inn: str | None = None
    for row_idx in range(4, (worksheet.max_row or 0) + 1):
        supplier_inn = normalize_inn(worksheet.cell(row_idx, 2).value)
        row_buyer_inn = normalize_inn(worksheet.cell(row_idx, 3).value)
        document_date = parse_excel_date(worksheet.cell(row_idx, 4).value)
        amount = parse_decimal(worksheet.cell(row_idx, 5).value)
        if supplier_inn is None and row_buyer_inn is None:
            if lines:
                break
            continue
        if supplier_inn is None or document_date is None or amount is None or amount <= 0:
            if lines:
                break
            continue
        if row_buyer_inn:
            buyer_inn = row_buyer_inn
        elif buyer_inn is None:
            continue
        lines.append((supplier_inn, buyer_inn or row_buyer_inn or "", document_date, amount))
    return lines


@dataclass
class OrderHit:
    order_id: int
    lead_id: int
    order_no: int
    status: str
    source_filename: str | None
    db_lines: int
    db_volume: Decimal
    new_lines: int
    new_volume: Decimal
    legacy_lines: int
    source: str
    storage_key: str | None


async def _fetch_bytes(storage_key: str) -> bytes | None:
    try:
        data, _ct = await get_file_storage().get_bytes(storage_key)
        return data
    except Exception as exc:
        _log(f"  WARN storage get failed key={storage_key!r}: {exc}")
        return None


async def _resolve_order_file(
    session: Any,
    *,
    order_id: int,
    chat_id: int | None,
    source_filename: str | None,
    source_message_id: int | None,
    source_attachment_index: int | None,
) -> tuple[bytes | None, str, str | None]:
    """Return (content, how, storage_key)."""
    if source_message_id is not None and source_attachment_index is not None:
        row = (
            await session.execute(
                text(
                    """
                    SELECT att->>'storage_key' AS storage_key,
                           coalesce(att->>'filename', att->>'name') AS name
                    FROM messages m
                    CROSS JOIN LATERAL jsonb_array_elements(
                      CASE WHEN jsonb_typeof(m.attachments)='array'
                           THEN m.attachments ELSE '[]'::jsonb END
                    ) WITH ORDINALITY AS t(att, ord)
                    WHERE m.id = :mid
                      AND (t.ord - 1) = :idx
                    LIMIT 1
                    """
                ),
                {"mid": source_message_id, "idx": int(source_attachment_index)},
            )
        ).mappings().first()
        if row and row["storage_key"]:
            data = await _fetch_bytes(str(row["storage_key"]))
            if data:
                return data, "source_attachment", str(row["storage_key"])

    fname = (source_filename or "").strip()
    if fname and chat_id is not None:
        row = (
            await session.execute(
                text(
                    """
                    SELECT att->>'storage_key' AS storage_key
                    FROM messages m
                    CROSS JOIN LATERAL jsonb_array_elements(
                      CASE WHEN jsonb_typeof(m.attachments)='array'
                           THEN m.attachments ELSE '[]'::jsonb END
                    ) att
                    WHERE m.chat_id = :cid
                      AND nullif(att->>'storage_key','') IS NOT NULL
                      AND lower(coalesce(att->>'filename', att->>'name', '')) = lower(:fname)
                    ORDER BY m.id DESC
                    LIMIT 1
                    """
                ),
                {"cid": chat_id, "fname": fname},
            )
        ).mappings().first()
        if row and row["storage_key"]:
            data = await _fetch_bytes(str(row["storage_key"]))
            if data:
                return data, "chat_filename", str(row["storage_key"])

        row = (
            await session.execute(
                text(
                    """
                    SELECT storage_key
                    FROM group_chat_files
                    WHERE chat_id = :cid
                      AND lower(original_name) = lower(:fname)
                    ORDER BY id DESC
                    LIMIT 1
                    """
                ),
                {"cid": chat_id, "fname": fname},
            )
        ).mappings().first()
        if row and row["storage_key"]:
            data = await _fetch_bytes(str(row["storage_key"]))
            if data:
                return data, "group_chat_files", str(row["storage_key"])

    _log(
        f"  no source file for order={order_id} "
        f"file={source_filename!r} msg={source_message_id} chat={chat_id}",
    )
    return None, "missing", None


async def _scan_orders(
    *,
    lead_id: int | None,
    only_suspects: bool,
    limit: int,
) -> list[OrderHit]:
    session_factory = get_session_factory()
    hits: list[OrderHit] = []
    missing = 0
    checked = 0
    async with session_factory() as session:
        rows = (
            await session.execute(
                text(
                    """
                    SELECT o.id AS order_id,
                           o.lead_id,
                           o.order_no,
                           o.status,
                           o.source_filename,
                           o.source_message_id,
                           o.source_attachment_index,
                           o.total_volume,
                           l.chat_id,
                           (SELECT count(*)::int
                              FROM lead_opt_order_lines ln
                             WHERE ln.order_id = o.id) AS db_lines,
                           (SELECT coalesce(sum(ln.amount), 0)
                              FROM lead_opt_order_lines ln
                             WHERE ln.order_id = o.id) AS db_volume
                    FROM lead_opt_orders o
                    JOIN leads l ON l.id = o.lead_id
                    WHERE o.deleted_at IS NULL
                      AND (:lead_id IS NULL OR o.lead_id = :lead_id)
                    ORDER BY o.id DESC
                    LIMIT :lim
                    """
                ),
                {"lead_id": lead_id, "lim": limit},
            )
        ).mappings().all()

        _log(f"orders to check: {len(rows)}")
        for row in rows:
            content, how, key = await _resolve_order_file(
                session,
                order_id=int(row["order_id"]),
                chat_id=int(row["chat_id"]) if row["chat_id"] is not None else None,
                source_filename=row["source_filename"],
                source_message_id=(
                    int(row["source_message_id"])
                    if row["source_message_id"] is not None
                    else None
                ),
                source_attachment_index=(
                    int(row["source_attachment_index"])
                    if row["source_attachment_index"] is not None
                    else None
                ),
            )
            if content is None:
                missing += 1
                continue
            checked += 1
            try:
                parsed = parse_application_workbook(content)
            except Exception as exc:
                _log(f"  parse fail order={row['order_id']}: {exc}")
                continue
            try:
                legacy = parse_legacy_truncated(content)
            except Exception:
                legacy = []

            db_lines = int(row["db_lines"] or 0)
            db_volume = Decimal(str(row["db_volume"] or 0))
            new_lines = len(parsed.lines)
            new_volume = sum((ln.amount for ln in parsed.lines), Decimal("0"))
            legacy_lines = len(legacy)

            suspect = new_lines > db_lines or (
                new_volume - db_volume > Decimal("0.05")
            ) or (legacy_lines < new_lines and legacy_lines == db_lines)

            if only_suspects and not suspect:
                continue

            hit = OrderHit(
                order_id=int(row["order_id"]),
                lead_id=int(row["lead_id"]),
                order_no=int(row["order_no"]),
                status=str(row["status"]),
                source_filename=row["source_filename"],
                db_lines=db_lines,
                db_volume=db_volume,
                new_lines=new_lines,
                new_volume=new_volume,
                legacy_lines=legacy_lines,
                source=how,
                storage_key=key,
            )
            hits.append(hit)
            flag = "SUSPECT" if suspect else "ok"
            _log(
                f"  [{flag}] order={hit.order_id} lead={hit.lead_id} no={hit.order_no} "
                f"file={hit.source_filename!r} via={hit.source} "
                f"db_lines={hit.db_lines} new={hit.new_lines} legacy={hit.legacy_lines} "
                f"db_vol={_money(hit.db_volume)} new_vol={_money(hit.new_volume)} "
                f"missing_lines={max(0, hit.new_lines - hit.db_lines)}",
            )

    _log(
        f"checked={checked} missing_file={missing} "
        f"reported={len(hits)} (only_suspects={only_suspects})",
    )
    return hits


async def _scan_files(*, limit: int) -> int:
    """Scan chat xlsx: legacy parse shorter than new parse → would have truncated."""
    session_factory = get_session_factory()
    async with session_factory() as session:
        rows = (
            await session.execute(
                text(
                    """
                    SELECT DISTINCT ON (att->>'storage_key')
                           att->>'storage_key' AS storage_key,
                           coalesce(att->>'filename', att->>'name', 'file.xlsx') AS name,
                           m.chat_id,
                           m.lead_id
                    FROM messages m
                    CROSS JOIN LATERAL jsonb_array_elements(
                      CASE WHEN jsonb_typeof(m.attachments)='array'
                           THEN m.attachments ELSE '[]'::jsonb END
                    ) att
                    WHERE nullif(att->>'storage_key','') IS NOT NULL
                      AND lower(coalesce(att->>'filename', att->>'name', ''))
                          ~ '\\.(xlsx|xlsm)$'
                    ORDER BY att->>'storage_key', m.id DESC
                    LIMIT :lim
                    """
                ),
                {"lim": limit},
            )
        ).mappings().all()

    _log(f"xlsx attachments to scan: {len(rows)}")
    suspects = 0
    for row in rows:
        key = str(row["storage_key"])
        data = await _fetch_bytes(key)
        if not data:
            continue
        try:
            parsed = parse_application_workbook(data)
            legacy = parse_legacy_truncated(data)
        except Exception:
            continue
        if len(legacy) >= len(parsed.lines):
            continue
        suspects += 1
        new_vol = sum((ln.amount for ln in parsed.lines), Decimal("0"))
        leg_vol = sum((ln[3] for ln in legacy), Decimal("0"))
        _log(
            f"  [FILE] chat={row['chat_id']} lead={row['lead_id']} "
            f"name={row['name']!r} legacy_lines={len(legacy)} new_lines={len(parsed.lines)} "
            f"legacy_vol={_money(leg_vol)} new_vol={_money(new_vol)} "
            f"key={key}",
        )
    _log(f"truncation-prone files: {suspects}")
    return 0


async def _amain(args: argparse.Namespace) -> int:
    if args.local_file:
        content = Path(args.local_file).read_bytes()
        parsed = parse_application_workbook(content)
        legacy = parse_legacy_truncated(content)
        _log(
            f"local {args.local_file}: new_lines={len(parsed.lines)} "
            f"legacy_lines={len(legacy)} "
            f"new_vol={_money(sum((ln.amount for ln in parsed.lines), Decimal('0')))}",
        )
        return 0
    if args.scan_files:
        return await _scan_files(limit=args.limit)
    hits = await _scan_orders(
        lead_id=args.lead_id,
        only_suspects=args.only_suspects,
        limit=args.limit,
    )
    suspects = [
        h
        for h in hits
        if h.new_lines > h.db_lines
        or (h.new_volume - h.db_volume > Decimal("0.05"))
        or (h.legacy_lines < h.new_lines and h.legacy_lines == h.db_lines)
    ]
    _log("")
    _log(f"SUSPECTS (truncated vs DB or legacy==db < new): {len(suspects)}")
    for h in suspects:
        _log(
            f"  lead={h.lead_id} order={h.order_id} no={h.order_no} "
            f"file={h.source_filename!r} "
            f"+{h.new_lines - h.db_lines} lines "
            f"+{_money(h.new_volume - h.db_volume)} ₽",
        )
    return 0


def main() -> None:
    try:
        sys.stdout.reconfigure(line_buffering=True)  # type: ignore[attr-defined]
    except Exception:
        pass
    p = argparse.ArgumentParser(description="Scan OPT orders truncated by old Excel parser")
    p.add_argument("--lead-id", type=int, default=None)
    p.add_argument("--only-suspects", action="store_true")
    p.add_argument("--scan-files", action="store_true", help="Scan chat xlsx for gap-truncation")
    p.add_argument("--local-file", type=str, default=None)
    p.add_argument("--limit", type=int, default=300)
    args = p.parse_args()
    raise SystemExit(asyncio.run(_amain(args)))


if __name__ == "__main__":
    main()
